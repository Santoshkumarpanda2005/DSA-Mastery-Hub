import openpyxl
import json

wb = openpyxl.load_workbook('LeetCode_Interview_Tracker.xlsx')
ws = wb.active

mapping = {}

prob_col_idx = None
link_col_idx = None

for cell in ws[1]:
    val = str(cell.value).lower()
    if 'problem' in val or 'title' in val:
        prob_col_idx = cell.column
    if 'link' in val or 'url' in val:
        link_col_idx = cell.column

if prob_col_idx and link_col_idx:
    for row in range(2, ws.max_row + 1):
        prob_cell = ws.cell(row=row, column=prob_col_idx)
        link_cell = ws.cell(row=row, column=link_col_idx)
        
        title = str(prob_cell.value).strip()
        
        link = ""
        if link_cell.hyperlink:
            link = link_cell.hyperlink.target
        elif link_cell.value and str(link_cell.value).startswith('http'):
            link = str(link_cell.value).strip()
            
        if title and link and title != 'None' and link != 'None':
            # Remove (Easy) (Medium) (Hard) from title if present, just mapping the base name
            mapping[title] = link

    with open('links.json', 'w') as f:
        json.dump(mapping, f, indent=2)
    print(f"Extracted {len(mapping)} links.")
else:
    print("Could not find columns")
